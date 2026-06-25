import sys
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read JSON input (passed via stdin or file)
if len(sys.argv) < 3:
    print("Usage: python generate_lme_excel.py <json_path> <output_xlsx_path> [week_header]")
    sys.exit(1)

json_path = sys.argv[1]
output_path = sys.argv[2]
week_header = sys.argv[3] if len(sys.argv) > 3 else None

with open(json_path, 'r', encoding='utf-8') as f:
    blocks = json.load(f)

# Find the block matching the week header
target_block = None
if week_header:
    for b in blocks:
        if b['header'] == week_header:
            target_block = b
            break
else:
    target_block = blocks[-1]  # Default to last block

if not target_block:
    print(f"Error: Block for week {week_header} not found.")
    sys.exit(1)

wb = Workbook()
ws = wb.active
ws.title = "TABELA LME"
ws.views.sheetView[0].showGridLines = True

# Helper definitions
font_family = "Segoe UI"
font_title = Font(name=font_family, size=18, bold=True, color="1F4E78")
font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
font_bold = Font(name=font_family, size=10, bold=True)
font_normal = Font(name=font_family, size=10)

fill_gray_header = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")

# Metal styling: color maps for header and data rows
metal_styles = {
    'cobre': {'hdr_fill': PatternFill(start_color="C65911", end_color="C65911", fill_type="solid"),
              'val_fill': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")},
    'zinco': {'hdr_fill': PatternFill(start_color="806000", end_color="806000", fill_type="solid"),
              'val_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")},
    'aluminio': {'hdr_fill': PatternFill(start_color="5B5B5B", end_color="5B5B5B", fill_type="solid"),
                 'val_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")},
    'chumbo': {'hdr_fill': PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid"),
               'val_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")},
    'estanho': {'hdr_fill': PatternFill(start_color="7F6000", end_color="7F6000", fill_type="solid"),
                'val_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")},
    'niquel': {'hdr_fill': PatternFill(start_color="333333", end_color="333333", fill_type="solid"),
               'val_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")},
    'dolar': {'hdr_fill': PatternFill(start_color="375623", end_color="375623", fill_type="solid"),
              'val_fill': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")}
}

thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

double_bottom = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='double', color='000000')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

# Title row
ws.merge_cells("B2:I2")
ws["B2"] = "APEXTECH METAIS"
ws["B2"].font = font_title
ws["B2"].alignment = align_left

# Main Table Headers (Row 4)
headers = ["DATA", "COBRE", "ZINCO", "ALUMÍNIO", "CHUMBO", "ESTANHO", "NÍQUEL", "DÓLAR"]
cols = ["B", "C", "D", "E", "F", "G", "H", "I"]
metal_keys = ["cobre", "zinco", "aluminio", "chumbo", "estanho", "niquel", "dolar"]

ws.row_dimensions[4].height = 24
for i, h in enumerate(headers):
    cell = ws[f"{cols[i]}4"]
    cell.value = h
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border
    if i == 0:
        cell.fill = fill_gray_header
    else:
        key = metal_keys[i - 1]
        cell.fill = metal_styles[key]['hdr_fill']

# Daily values (Rows 5 to 9)
for idx, d in enumerate(target_block['days']):
    r = 5 + idx
    ws.row_dimensions[r].height = 20
    
    # Date cell
    cell_date = ws[f"B{r}"]
    cell_date.value = d['data']
    cell_date.font = font_normal
    cell_date.alignment = align_center
    cell_date.border = thin_border
    
    # Metal values
    for m_idx, m in enumerate(metal_keys):
        cell_val = ws[f"{cols[m_idx + 1]}{r}"]
        val = d[m]
        if val == 'feriado':
            cell_val.value = 'feriado'
            cell_val.font = Font(name=font_family, size=10, italic=True, color="555555")
        elif val is not None:
            cell_val.value = val
            cell_val.font = font_normal
            if m == 'dolar':
                cell_val.number_format = '0.0000'
            else:
                cell_val.number_format = '#,##0.00'
        cell_val.alignment = align_center
        cell_val.border = thin_border

# Computed Rows (Row 10 to 16)
comp = target_block.get('computed', {})
comp_rows = [
    ("MEDIA SEMANAL", "MEDIA SEMANAL", PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"), '#,##0.00', '0.0000'),
    ("100% LME", "100% LME", PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), 'R$ #,##0.000', None),
    ("SEMANA ANTERIOR", "SEMANA ANTERIOR", PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"), 'R$ #,##0.000', 'R$ #,##0.0000'),
    ("FECHAMENTO % ( SEMANA ANTERIOR )", "FECHAMENTO % ( SEMANA ANTERIOR )", PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), '0.000%', '0.000%'),
    ("OSCILAÇÃO %", "OSCILAÇÃO %", PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"), '0.000%', '0.000%'),
    ("OSCILAÇÃO R$", "OSCILAÇÃO R$", PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"), 'R$ #,##0.0000', 'R$ #,##0.0000'),
    ("MEDIA MENSAL", "MEDIA MENSAL", PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), 'R$ #,##0.00', 'R$ #,##0.0000')
]

for row_offset, (label, json_key, fill, num_fmt, dol_fmt) in enumerate(comp_rows):
    r = 10 + row_offset
    ws.row_dimensions[r].height = 22
    
    cell_lbl = ws[f"B{r}"]
    cell_lbl.value = label
    cell_lbl.font = font_bold
    cell_lbl.alignment = align_center
    cell_lbl.border = thin_border
    cell_lbl.fill = fill
    
    row_vals = comp.get(json_key, {})
    for m_idx, m in enumerate(metal_keys):
        cell_val = ws[f"{cols[m_idx + 1]}{r}"]
        val = row_vals.get(m)
        if val is not None and not isinstance(val, str):
            cell_val.value = val
            cell_val.font = font_bold
            cell_val.number_format = dol_fmt if (m == 'dolar' and dol_fmt) else num_fmt
        elif isinstance(val, str):
            cell_val.value = val
            cell_val.font = font_bold
            
        cell_val.alignment = align_center
        cell_val.border = thin_border
        cell_val.fill = fill

# Summary Table Header (Row 18)
ws.row_dimensions[18].height = 24
for i, h in enumerate(["TIPO", "COBRE", "ZINCO", "ALUMÍNIO", "CHUMBO", "ESTANHO", "NÍQUEL", "DÓLAR"]):
    cell = ws[f"{cols[i]}18"]
    cell.value = h
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border
    if i == 0:
        cell.fill = fill_gray_header
    else:
        key = metal_keys[i - 1]
        cell.fill = metal_styles[key]['hdr_fill']

# Summary Values (Rows 19 to 21)
# 19: SEMANA ANTERIOR
# 20: LME ATUAL
# 21: Osilacao
summary_rows = [
    ("SEMANA ANTERIOR", "SEMANA ANTERIOR", 'R$ #,##0.00', 'R$ #,##0.0000'),
    ("LME ATUAL", "100% LME", 'R$ #,##0.00', 'R$ #,##0.0000'),
]

for row_offset, (label, json_key, num_fmt, dol_fmt) in enumerate(summary_rows):
    r = 19 + row_offset
    ws.row_dimensions[r].height = 20
    cell_lbl = ws[f"B{r}"]
    cell_lbl.value = label
    cell_lbl.font = font_bold
    cell_lbl.alignment = align_center
    cell_lbl.border = thin_border
    
    row_vals = comp.get(json_key, {})
    for m_idx, m in enumerate(metal_keys):
        cell_val = ws[f"{cols[m_idx + 1]}{r}"]
        val = row_vals.get(m)
        if val is not None and not isinstance(val, str):
            cell_val.value = val
            cell_val.font = font_normal
            cell_val.number_format = dol_fmt if m == 'dolar' else num_fmt
        elif isinstance(val, str):
            cell_val.value = val
            cell_val.font = font_normal
            
        cell_val.alignment = align_center
        cell_val.border = thin_border

# 21: Osilacao
r = 21
ws.row_dimensions[r].height = 22
cell_lbl = ws[f"B{r}"]
cell_lbl.value = "Osilacao"
cell_lbl.font = Font(name=font_family, size=10, bold=True, italic=True)
cell_lbl.alignment = align_left
cell_lbl.border = thin_border

osc_vals = comp.get("OSCILAÇÃO R$", {})
for m_idx, m in enumerate(metal_keys):
    cell_val = ws[f"{cols[m_idx + 1]}{r}"]
    val = osc_vals.get(m)
    if val is not None and not isinstance(val, str):
        # Format like: Up arrow + R$ 0,0586 or Down arrow + R$ -1,4462
        arrow = "▲ " if val >= 0 else "▼ "
        cell_val.value = arrow + ("R$ " if m != 'dolar' else "") + f"{val:+.4f}"
        
        # Color text green/red
        text_color = "385723" if val >= 0 else "C65911"
        cell_val.font = Font(name=font_family, size=10, bold=True, color=text_color)
    else:
        cell_val.value = ""
        cell_val.font = font_bold
        
    cell_val.alignment = align_center
    cell_val.border = thin_border

# Autofit column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    if col_letter == 'A':
        ws.column_dimensions[col_letter].width = 3
        continue
    for cell in col:
        # Ignore title row merged cell and long label cells to prevent giant column widths
        if cell.row == 2:
            continue
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

# Specific column overrides
ws.column_dimensions['B'].width = 34  # Header/labels column

# Save workbook
wb.save(output_path)
print(f"Excel report saved successfully to {output_path}")
sys.exit(0)
